import os
import pandas as pd
import json
from datetime import datetime
from openpyxl.utils import column_index_from_string
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, numbers, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.worksheet import Worksheet

months = ['janv', 'fev', 'mars', 'avr', 'mai', 'juin',
          'juill', 'aout', 'sept', 'oct', 'nov', 'dec']


def create_new_budget(year):
    current_dir = os.path.dirname(os.path.realpath(__file__))
    refdep_json_path = os.path.join(current_dir, "ressources", "refdep.json")
    reference_budget_path = os.path.join(current_dir, "ressources", "budget.xlsx")

    with open(refdep_json_path, "r", encoding="utf-8") as f:
        refdep_data = json.load(f)

    intitule_to_code = {
        entry["intitule"].strip().lower(): code
        for code, entry in refdep_data.items()
        if "intitule" in entry
    }

    wb_source = load_workbook(reference_budget_path, data_only=True)
    if "actuel" not in wb_source.sheetnames:
        raise ValueError("Sheet 'actuel' not found in the source budget file.")
    ws_source = wb_source["actuel"]

    categories = []
    for row in ws_source.iter_rows(min_row=5, min_col=1, max_col=3, values_only=True):
        intitule, credit, debit = row
        if intitule and str(intitule).strip().lower() != "salaire":
            intitule_clean = str(intitule).strip()
            code = intitule_to_code.get(intitule_clean.lower(), "UNKNOWN")
            categories.append({
                "code": code,
                "intitule": intitule_clean,
                "prevu": debit if debit is not None else 0
            })

    wb_new = Workbook()
    wb_new.remove(wb_new.active)

    headers = ["Code", "Intitulé", "Prévu", "Réel", "Conclusion", "Reste"]
    operation_headers = ["Date opération", "Type", "Libellé", "Crédit", "Débit", "Solde"]

    for month in months:
        ws = wb_new.create_sheet(title=month)

        # Write header row
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write operation headers from column H (index 8)
        for op_idx, header in enumerate(operation_headers, start=8):
            ws.cell(row=1, column=op_idx, value=header)

        # Write data rows
        for row_idx, entry in enumerate(categories, start=5):
            ws.cell(row=row_idx, column=1, value=entry["code"])
            ws.cell(row=row_idx, column=2, value=entry["intitule"])

            # Column C (Prévu)
            cell_prevu = ws.cell(row=row_idx, column=3, value=entry["prevu"])
            cell_prevu.number_format = '# ##0 €'

            # Column D (Réel)
            formula_reel = f'=IFERROR(SUMIF($I:$I, A{row_idx}, $L:$L), 0)'
            cell_reel = ws.cell(row=row_idx, column=4, value=formula_reel)
            cell_reel.number_format = '# ##0.00 €'

            # Column E (Conclusion)
            cell_conclusion = ws.cell(row=row_idx, column=5, value=f'=IF(F{row_idx}>=0,"OK","PB")')
            cell_conclusion.alignment = Alignment(horizontal="center")

            # Column F (Reste)
            formula_reste = (
                f'=C{row_idx} - (IFERROR(SUMIF($I:$I, A{row_idx}, $L:$L), 0) - '
                f'IFERROR(SUMIF($I:$I, A{row_idx}, $K:$K), 0))'
            )
            cell_reste = ws.cell(row=row_idx, column=6, value=formula_reste)
            cell_reste.number_format = '# ##0.00 €'

        # Apply conditional formatting to Conclusion (col E)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        font_ok = Font(color="006100")
        font_pb = Font(color="9C0006")

        ws.conditional_formatting.add(
            f"E5:E{4 + len(categories)}",
            FormulaRule(formula=[f'E5="OK"'], fill=green_fill, font=font_ok)
        )
        ws.conditional_formatting.add(
            f"E5:E{4 + len(categories)}",
            FormulaRule(formula=[f'E5="PB"'], fill=red_fill, font=font_pb)
        )

        # Column widths
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 1
        fill_grey = PatternFill(start_color="4D4D4D", end_color="4D4D4D", fill_type="solid")
        for row in range(1, 100):  # Ajuste le nombre de lignes selon ton besoin
            ws.cell(row=row, column=7).fill = fill_grey  # column 7 = G
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 5
        ws.column_dimensions["J"].width = 80
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 10

    # Add stats sheet
    generate_stats_sheet(wb_new, categories)

    # Save file
    output_path = f"{year}.xlsx"
    wb_new.save(output_path)

    print(f"✅ Budget file created: {output_path}")


def import_month():
    """
    Ask the user which month to import, load the corresponding CSV file,
    and overwrite the data in the corresponding sheet of the Excel file
    starting from row 2 and column H (columns H to M will be cleared).
    """
    year = datetime.now().year
    months_map = {
        "1": "janv", "2": "fev", "3": "mars", "4": "avr",
        "5": "mai", "6": "juin", "7": "juill", "8": "aout",
        "9": "sept", "10": "oct", "11": "nov", "12": "dec"
    }

    # Display month selection
    print("📅 Available months to import: " + " | ".join([f"{num}. {name}" for num, name in months_map.items()]))
    choice = input("➡️  Enter the number of the month to import: ").strip()
    selected_month = months_map.get(choice)

    if not selected_month:
        print("❌ Invalid choice. Please enter a number between 1 and 12.")
        return

    csv_filename = f"{selected_month}.csv"
    if not os.path.exists(csv_filename):
        print(f"❌ File not found: {csv_filename}")
        return

    try:
        # Load the CSV file (UTF-16 with tab separator)
        df = pd.read_csv(csv_filename, sep='\t', encoding='utf-16')
        df['Montant'] = (
            df['Montant'].astype(str)
            .str.replace(',', '.', regex=False)
            .str.replace(' ', '', regex=False)
            .astype(float)
        )
    except Exception as e:
        print(f"❌ Error reading or processing the file: {e}")
        return

    excel_filename = f"{year}.xlsx"
    if not os.path.exists(excel_filename):
        print(f"❌ Excel budget file '{excel_filename}' not found.")
        return

    wb = load_workbook(excel_filename)
    if selected_month not in wb.sheetnames:
        print(f"❌ Sheet '{selected_month}' not found in the workbook.")
        return

    ws = wb[selected_month]

    # Clear previous data from row 2 down, columns H to M
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=column_index_from_string("H"),
                            max_col=column_index_from_string("M")):
        for cell in row:
            cell.value = None

    # Mapping columns
    start_row = 2
    col_map = {
        "Date": "H",
        "Libellé": "J",
        "Débit": "L",
        "Crédit": "K"
    }

    for i, (_, row) in enumerate(df.iterrows(), start=start_row):
        ws[f"{col_map['Date']}{i}"] = row.get("Date", "")
        ws[f"{col_map['Libellé']}{i}"] = row.get("Libellé", "")
        montant = row.get("Montant", 0.0)
        if montant < 0:
            ws[f"{col_map['Débit']}{i}"] = abs(montant)
        else:
            ws[f"{col_map['Crédit']}{i}"] = montant

    fill_category_codes(ws)
    wb.save(excel_filename)

    print(f"✅ Imported {len(df)} operations into '{excel_filename}', sheet '{selected_month}'.")


def assign_categories():
    """
    Ask the user for a month, then fill category codes (column I)
    in the corresponding sheet of the Excel budget file.
    """
    year = datetime.now().year
    excel_filename = f"{year}.xlsx"

    if not os.path.exists(excel_filename):
        print(f"❌ Excel budget file '{excel_filename}' not found.")
        return

    months_map = {
        "1": "janv", "2": "fev", "3": "mars", "4": "avr",
        "5": "mai", "6": "juin", "7": "juill", "8": "aout",
        "9": "sept", "10": "oct", "11": "nov", "12": "dec"
    }

    print("📅 Available months: " + " | ".join([f"{num}. {name}" for num, name in months_map.items()]))
    choice = input("➡️  Enter the number of the month to assign categories: ").strip()
    selected_month = months_map.get(choice)

    if not selected_month:
        print("❌ Invalid choice. Please enter a number between 1 and 12.")
        return

    wb = load_workbook(excel_filename)
    if selected_month not in wb.sheetnames:
        print(f"❌ Sheet '{selected_month}' not found.")
        return

    ws = wb[selected_month]
    fill_category_codes(ws)
    wb.save(excel_filename)
    print(f"✅ Categories assigned for '{selected_month}' in '{excel_filename}'.")


def fill_category_codes(sheet: Worksheet, refdep_path="refdep.json", start_row=2):
    # Always resolve relative to current Python script's location
    current_dir = os.path.dirname(os.path.realpath(__file__))
    refdep_path = os.path.join(current_dir, "ressources", refdep_path)
    """
    Fill column I in the given worksheet based on matching keywords in column J (Libellé),
    using the reference category codes from the refdep JSON file.

    Parameters:
    - sheet (Worksheet): The openpyxl worksheet object to modify.
    - refdep_path (str): Path to the JSON file containing the mapping of keywords to category codes.
    - start_row (int): Row number to start processing from (default is 2).
    """
    # Load category reference JSON
    with open(refdep_path, "r", encoding="utf-8") as f:
        refdep_data = json.load(f)

    # Build a dictionary mapping keyword (lowercased) → code
    keyword_to_code = {}
    for code, info in refdep_data.items():
        if "intitule" in info and "valeurs" in info:
            for keyword in info["valeurs"]:
                keyword_lower = keyword.strip().lower()
                keyword_to_code[keyword_lower] = code

    # Process each row in column J (Libellé), write code to column I
    for row in sheet.iter_rows(min_row=start_row, min_col=10, max_col=10):  # Column J
        libelle_cell = row[0]
        if libelle_cell.value:
            libelle_text = str(libelle_cell.value).lower()
            for keyword, code in keyword_to_code.items():
                if keyword in libelle_text:
                    sheet.cell(row=libelle_cell.row, column=9, value=code)  # Column I
                    break  # Stop at first match

    print("✅ Column I (category codes) filled based on keywords in column J.")


def compare_budget_categories():
    """
    Compare budget categories from an Excel file with the 'intitule' values in a reference JSON file.

    Parameters:
    - budget_file_path (str): Path to the Excel budget file.
    - refdep_json_path (str): Path to the reference JSON file.

    Returns:
    - missing (list): List of budget categories not found in the 'intitule' fields of the JSON (case-insensitive).
    """
    refdep_json_path = os.path.join("./ressources", "refdep.json")
    budget_file_path = os.path.join("./ressources", "budget.xlsx")
    # Load the JSON file
    with open(refdep_json_path, "r", encoding="utf-8") as f:
        refdep_json = json.load(f)

    # Extract all 'intitule' values from the JSON (case-insensitive)
    known_intitules = set(
        entry["intitule"].strip().lower()
        for entry in refdep_json.values()
        if "intitule" in entry
    )

    # Load the Excel workbook and access the 'actuel' sheet
    wb = load_workbook(budget_file_path, data_only=True)
    if "actuel" not in wb.sheetnames:
        raise ValueError("Sheet 'actuel' not found in the Excel file.")
    ws = wb["actuel"]

    # Extract budget category names from column A starting at row 5, excluding 'salaire'
    categories = []
    for row in ws.iter_rows(min_row=5, max_col=1, values_only=True):
        value = row[0]
        if value and str(value).strip().lower() != "salaire":
            categories.append(str(value).strip())

    # Compare categories to known 'intitule' values
    missing = [cat for cat in categories if cat.lower() not in known_intitules]

    return missing


def generate_stats_sheet(wb, categories):
    """
    Create a 'stat' sheet summarizing monthly data for each category.
    For each month and category, it displays: Réel.
    Then it calculates: Total annuel and Monthly average (only on non-zero values).
    """

    ws = wb.create_sheet("stat")

    # Header titles
    ws.cell(row=1, column=1, value="Code")
    ws.cell(row=1, column=2, value="Catégorie")

    for i, month in enumerate(months):
        col = i + 3
        cell = ws.cell(row=1, column=col, value=month)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Additional headers
    ws.cell(row=1, column=15, value="Total an").font = Font(bold=True)
    ws.cell(row=1, column=15).alignment = Alignment(horizontal="center")

    ws.cell(row=1, column=16, value="Catégorie").font = Font(bold=True)
    ws.cell(row=1, column=16).alignment = Alignment(horizontal="center")

    ws.cell(row=1, column=17, value="Moyenne").font = Font(bold=True)
    ws.cell(row=1, column=17).alignment = Alignment(horizontal="center")

    # First two cells A1/A2 and B1/B2 in bold
    ws["A1"].font = Font(bold=True)
    ws["A2"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    # Write rows per category
    for row_idx, entry in enumerate(categories, start=2):
        ws.cell(row=row_idx, column=1, value=entry["code"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=entry["intitule"]).font = Font(bold=True)
        ws.cell(row=row_idx, column=16, value=entry["intitule"]).font = Font(bold=True)

        for i, month in enumerate(months):
            col = i + 3
            base_row = row_idx + 3
            formula = f"='{month}'!D{base_row}"
            ws.cell(row=row_idx, column=col, value=formula)

        # Total annuel
        ws.cell(row=row_idx, column=15, value=f"=SUM(C{row_idx}:N{row_idx})")

        # Moyenne avec format numérique
        formula = f'=IF(COUNTIF(C{row_idx}:N{row_idx}, ">0")=0, 0, O{row_idx} / COUNTIF(C{row_idx}:N{row_idx}, ">0"))'
        moyenne_cell = ws.cell(row=row_idx, column=17,
                               value=formula)
        moyenne_cell.number_format = numbers.FORMAT_NUMBER_00

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["O"].width = 10
    ws.column_dimensions["P"].width = 25
    ws.column_dimensions["Q"].width = 12
